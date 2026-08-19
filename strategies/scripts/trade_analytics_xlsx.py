#!/usr/bin/env python3
"""trade_analytics_xlsx.py — build log/trade_analytics.xlsx from the live trade journal.

DIFFERENT FROM straddle-income-growth-analysis.xlsx. That one is a pure CALCULATOR: pick a
hypothetical monthly return and see where compounding leads. This one is grounded in what
actually happened — every number on every sheet derives from log/trade_journal.csv — and its
Projection tab is anchored on the MEASURED pace, with tuning knobs on top.

Design rules:
  * Data is written as values; every other sheet is LIVE FORMULAS over it, so pasting extra
    rows into Data updates Summary, Monthly, the charts and the Projection anchor.
  * Formulas span rows 2:400 so the workbook keeps working as the journal grows.
  * The uncertainty travels WITH the headline. At n=7 a short-vol strategy can look like edge
    purely from a short tail, so t-stat, 95% bands and profit concentration sit next to the
    return, not in a footnote.
  * Return is on margin_blocked (broker-actual), never max_risk_defined — they differ ~4.6x.

Usage:  python trade_analytics_xlsx.py
Env:    TRADE_JOURNAL_CSV, ANALYTICS_XLSX
"""
import csv
import datetime as dt
import os
import sys

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.chart.marker import Marker
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.normpath(os.path.join(SCRIPT_DIR, "..", ".."))
JOURNAL = os.getenv("TRADE_JOURNAL_CSV", os.path.join(REPO, "log", "trade_journal.csv"))
OUT = os.getenv("ANALYTICS_XLSX", os.path.join(REPO, "log", "trade_analytics.xlsx"))

# Cash in the account BEFORE the first traded day. Together with SUM(net_pnl) this reproduces
# the live balance, so the Projection's starting corpus never needs hand-editing.
#
# NOT hardcoded on purpose: the mkds-openalgo remote is a PUBLIC fork of marketcalls/openalgo,
# and an opening balance is an account size, not a strategy detail. Set it via the OPENING_CASH
# env var or log/opening_cash.txt (both untracked). With neither, the corpus falls back to the
# cumulative P&L alone and the sheet says so, rather than silently projecting off a wrong base.
#
# How to derive it: take any broker cash reading whose settlement state you know, and subtract
# the journal's cumulative net up to that point. Do it from TWO readings on different days and
# check they agree -- ours agreed to 5 paise across 8 trading days, which is also what settles
# backlog #6: an identity that closes that tightly over the whole run says the fill-derived
# 08-10 net was right and the lower figure seen that day was a mid-settlement snapshot, so
# there was never a Rs50.45 discrepancy to chase.
def _opening_cash():
    v = os.getenv("OPENING_CASH")
    if not v:
        f = os.path.join(REPO, "log", "opening_cash.txt")
        if os.path.exists(f):
            v = open(f).read().strip()
    try:
        return float(str(v).replace(",", ""))
    except (TypeError, ValueError):
        return 0.0


OPENING_CASH = _opening_cash()

LAST = 400                     # formula horizon — the journal can grow to here untouched
PROJ_MONTHS = 60               # physical projection rows; the Months input blanks the rest

INR = '[>=10000000]₹##\\,##\\,##\\,##0;[>=100000]₹##\\,##\\,##0;₹##,##0'
INR2 = '₹#,##0.00'
PCT3 = '0.000%'
PCT2 = '0.00%'
DATEF = 'dd-mmm-yy'

H1 = Font(bold=True, size=14, color="1F3864")
H2 = Font(bold=True, size=11, color="FFFFFF")
LBL = Font(bold=True)
NOTE = Font(italic=True, size=9, color="666666")
BIG = Font(bold=True, size=12, color="1F3864")
IN_FILL = PatternFill("solid", fgColor="FFF2CC")
HD_FILL = PatternFill("solid", fgColor="4472C4")
SEC_FILL = PatternFill("solid", fgColor="D9E1F2")
thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
GOOD = PatternFill("solid", fgColor="C6EFCE")
BAD = PatternFill("solid", fgColor="FFC7CE")


def load():
    if not os.path.exists(JOURNAL):
        sys.exit(f"journal not found: {JOURNAL}")
    return list(csv.DictReader(open(JOURNAL)))


def num(v):
    try:
        return float(str(v).replace(",", ""))
    except (TypeError, ValueError):
        return None


rows = load()
COLS = list(rows[0].keys())
NROWS = len(rows)
wb = Workbook()

# ════════════════════════════════════════════════════════ 1. DATA
ws = wb.active
ws.title = "Data"
ws.freeze_panes = "C2"
for i, c in enumerate(COLS, start=1):
    cell = ws.cell(row=1, column=i, value=c)
    cell.font, cell.fill, cell.border = H2, HD_FILL, BOX
    cell.alignment = Alignment(wrap_text=True, vertical="center")
DATEC = {"date", "expiry"}
INTC = {"dte", "lots", "qty", "orb_range", "atm_strike", "wing_width", "breach_lo", "breach_hi",
        "spot_exit", "n_orders", "n_fills", "mfe", "mae", "gross_premium", "hedge_cost",
        "premium_collected", "max_risk_defined"}
MONEY = {"margin_blocked", "gross_pnl", "charges", "net_pnl", "slip_entry", "slip_exit",
         "tg_target_net", "tg_stop_net"}
for r, rec in enumerate(rows, start=2):
    for i, c in enumerate(COLS, start=1):
        v = rec[c]
        cell = ws.cell(row=r, column=i)
        if c in DATEC and v:
            try:
                cell.value = dt.datetime.strptime(v, "%Y-%m-%d") if c == "date" \
                    else dt.datetime.strptime(v, "%d-%b-%y")
                cell.number_format = DATEF
                continue
            except ValueError:
                pass
        n = num(v)
        if n is not None and v != "":
            cell.value = n
            if c in MONEY:
                cell.number_format = INR2
            elif c in INTC:
                cell.number_format = "#,##0"
            elif "pct" in c:
                cell.number_format = '0.000'
        else:
            cell.value = v or None
# month key for the Monthly sheet — added by us, not part of the journal
mk = len(COLS) + 1
ws.cell(row=1, column=mk, value="month_key").font = H2
ws.cell(row=1, column=mk).fill = HD_FILL
for r in range(2, NROWS + 2):
    ws.cell(row=r, column=mk, value=f'=IF($A{r}="","",TEXT($A{r},"yyyy-mm"))')
ws.add_table(Table(displayName="Journal",
                   ref=f"A1:{get_column_letter(mk)}{NROWS + 1}",
                   tableStyleInfo=TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)))
widths = {"date": 11, "weekday": 8, "series_code": 14, "expiry": 11, "exit_reason": 22,
          "confidence": 11, "notes": 46, "margin_blocked": 14, "net_pnl": 12, "gross_pnl": 12}
for i, c in enumerate(COLS, start=1):
    ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 11)
ws.column_dimensions[get_column_letter(mk)].width = 11
# colour the P&L column
npc = COLS.index("net_pnl") + 1
L = get_column_letter(npc)
ws.conditional_formatting.add(f"{L}2:{L}{LAST}",
                              CellIsRule(operator="greaterThan", formula=["0"], fill=GOOD))
ws.conditional_formatting.add(f"{L}2:{L}{LAST}",
                              CellIsRule(operator="lessThan", formula=["0"], fill=BAD))

C = {c: get_column_letter(i) for i, c in enumerate(COLS, start=1)}
D = "Data!"
def rng(col):
    return f"{D}${C[col]}$2:${C[col]}${LAST}"
MONTHK = f"{D}${get_column_letter(mk)}$2:${get_column_letter(mk)}${LAST}"

# ════════════════════════════════════════════════════════ 2. LIVE STATUS
s = wb.create_sheet("Live Status")
s.sheet_view.showGridLines = False
s["A1"] = "Live Trading Status — everything below is a formula over the Data sheet"
s["A1"].font = H1
s["A2"] = '=" through "&TEXT(MAX(' + rng("date") + '),"dd-mmm-yyyy")&"  ·  "&COUNT(' + rng("date") + ')&" traded days"'
s["A2"].font = Font(italic=True, color="1F3864")

def sec(row, text, width=6):
    s.cell(row=row, column=1, value=text).font = Font(bold=True, color="1F3864")
    for c in range(1, width + 1):
        s.cell(row=row, column=c).fill = SEC_FILL

def kv(row, label, formula, fmt=None, note=None, big=False):
    s.cell(row=row, column=1, value=label).font = LBL
    c = s.cell(row=row, column=2, value=formula)
    if fmt:
        c.number_format = fmt
    if big:
        c.font = BIG
    if note:
        s.cell(row=row, column=3, value=note).font = NOTE

sec(4, "RESULT")
kv(5, "Traded days", f"=COUNT({rng('date')})", "0")
kv(6, "Gross P&L", f"=SUM({rng('gross_pnl')})", INR2)
kv(7, "Zerodha charges", f"=-SUM({rng('charges')})", INR2)
kv(8, "NET P&L", f"=SUM({rng('net_pnl')})", INR2, "what actually landed", big=True)
kv(9, "Slippage (entry+exit)", f"=SUM({rng('slip_entry')})+SUM({rng('slip_exit')})", INR2,
   "already inside gross — shown to size the cost, not to subtract again")
kv(10, "Charge drag", f"=IF(SUM({rng('gross_pnl')})<=0,\"n/a\",-B7/SUM({rng('gross_pnl')}))", PCT2,
   "charges as a share of gross")
kv(11, "Slippage vs net", f"=IF(B8<=0,\"n/a\",B9/B8)", PCT2, "how much of the profit execution costs")

sec(13, "RETURN — on margin_blocked (broker-actual), never max_risk_defined")
kv(14, "Average margin blocked", f"=AVERAGE({rng('margin_blocked')})", INR2)
kv(15, "Margin range", f'="₹"&TEXT(MIN({rng("margin_blocked")}),"#,##0")&"  –  ₹"&TEXT(MAX({rng("margin_blocked")}),"#,##0")')
kv(16, "Margin per lot", f"=B14/AVERAGE({rng('lots')})", INR2, "measured — the Projection uses this")
kv(17, "Period return on avg margin", "=B8/B14", PCT2, "cumulative, whole live era", big=True)
kv(18, "Mean daily return on margin", f"=AVERAGE({rng('roi_on_margin_pct')})/100", PCT3)
kv(19, "Median daily", f"=MEDIAN({rng('roi_on_margin_pct')})/100", PCT3,
   "median vs mean shows how much a short tail is carrying it")
kv(20, "Best day / Worst day",
   f'=TEXT(MAX({rng("roi_on_margin_pct")})/100,"+0.000%")&"   /   "&TEXT(MIN({rng("roi_on_margin_pct")})/100,"+0.000%")')
kv(21, "Naive monthly run-rate", "=B17/B5*20", PCT2, "period return ÷ days × 20. Arithmetic, NOT a forecast")

sec(23, "CONSISTENCY")
kv(24, "Winning days", f'=COUNTIF({rng("net_pnl")},">0")&" / "&B5')
kv(25, "Win rate", f'=COUNTIF({rng("net_pnl")},">0")/B5', PCT2)
kv(26, "Avg win", f'=IFERROR(AVERAGEIF({rng("net_pnl")},">0"),"—")', INR2)
kv(27, "Avg loss", f'=IFERROR(AVERAGEIF({rng("net_pnl")},"<0"),"—")', INR2)
kv(28, "Expectancy per day", "=B8/B5", INR2)
kv(29, "Breached days", f'=COUNTIF({rng("breached")},"Y")&" / "&B5', None, "0.55% directional stop fired")
kv(30, "/stradexit fired", f'=COUNTIF({rng("tg_fired")},"Y")&" of "&COUNTIF({rng("tg_fired")},"<>")&" days armed"')

sec(32, "IS IT MEASURED YET?  ← read this before trusting anything above")
kv(33, "Std deviation, daily", f"=STDEV({rng('roi_on_margin_pct')})/100", PCT3)
kv(34, "Standard error", f"=B33/SQRT(COUNT({rng('roi_on_margin_pct')}))", PCT3)
kv(35, "t-statistic", "=B18/B34", "0.00",
   "≈2 is the usual bar for 'probably not luck'")
s["C36"] = '=IF(ABS(B35)<2,"NOT yet distinguishable from zero — treat the return as unproven","clears the usual t≈2 bar")'
s["C36"].font = Font(bold=True, color="C00000")
kv(37, "95% band, daily", '="from "&TEXT(B18-2*B34,"+0.000%")&"  to  "&TEXT(B18+2*B34,"+0.000%")')
kv(38, "95% band, monthly", '="from "&TEXT((B18-2*B34)*20,"+0.0%")&"  to  "&TEXT((B18+2*B34)*20,"+0.0%")',
   None, "×20 trading days — the width IS the point")
kv(39, "Top-2 days share of net",
   f'=IFERROR((LARGE({rng("net_pnl")},1)+LARGE({rng("net_pnl")},2))/B8,"—")', PCT2,
   "if this is near or above 100%, the result is a tail not a trend")
kv(40, "Net without top-2 days",
   f'=B8-LARGE({rng("net_pnl")},1)-LARGE({rng("net_pnl")},2)', INR2)
kv(41, "Rows not fill-verified", f'=COUNTIF({rng("confidence")},"<>high")&" of "&B5',
   None, "medium/low rows should be excluded from fill-level analysis")
for col, w in (("A", 32), ("B", 20), ("C", 70)):
    s.column_dimensions[col].width = w

# ════════════════════════════════════════════════════════ 3. MONTHLY
m = wb.create_sheet("Monthly")
m.sheet_view.showGridLines = False
m["A1"] = "Monthly Status"
m["A1"].font = H1
m["A2"] = "One row per calendar month, all formulas — new months populate themselves as Data grows."
m["A2"].font = NOTE
MH = ["Month", "Days", "Gross", "Charges", "Slippage", "NET", "Avg margin", "Return on margin",
      "Wins", "Win rate", "Best day", "Worst day", "Breached", "stradexit fired"]
for i, h in enumerate(MH, start=1):
    c = m.cell(row=4, column=i, value=h)
    c.font, c.fill, c.border = H2, HD_FILL, BOX
    c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    m.column_dimensions[get_column_letter(i)].width = 14 if i > 1 else 11
months = sorted({r["date"][:7] for r in rows if r.get("date")})
for i in range(24):                       # 24 physical month rows
    r = 5 + i
    mv = months[i] if i < len(months) else ""
    m.cell(row=r, column=1, value=mv or None)
    g = f'IF($A{r}="","",'
    f = {
        2: f'{g}COUNTIFS({MONTHK},$A{r}))',
        3: f'{g}SUMIFS({rng("gross_pnl")},{MONTHK},$A{r}))',
        4: f'{g}-SUMIFS({rng("charges")},{MONTHK},$A{r}))',
        5: f'{g}SUMIFS({rng("slip_entry")},{MONTHK},$A{r})+SUMIFS({rng("slip_exit")},{MONTHK},$A{r}))',
        6: f'{g}SUMIFS({rng("net_pnl")},{MONTHK},$A{r}))',
        7: f'{g}IFERROR(AVERAGEIFS({rng("margin_blocked")},{MONTHK},$A{r}),""))',
        8: f'{g}IFERROR($F{r}/$G{r},""))',
        9: f'{g}COUNTIFS({MONTHK},$A{r},{rng("net_pnl")},">0"))',
        10: f'{g}IFERROR($I{r}/$B{r},""))',
        11: f'{g}IFERROR(MAXIFS({rng("net_pnl")},{MONTHK},$A{r}),""))',
        12: f'{g}IFERROR(MINIFS({rng("net_pnl")},{MONTHK},$A{r}),""))',
        13: f'{g}COUNTIFS({MONTHK},$A{r},{rng("breached")},"Y"))',
        14: f'{g}COUNTIFS({MONTHK},$A{r},{rng("tg_fired")},"Y"))',
    }
    for col, formula in f.items():
        c = m.cell(row=r, column=col, value="=" + formula)
        c.number_format = {3: INR2, 4: INR2, 5: INR2, 6: INR2, 7: INR2,
                           8: PCT2, 10: PCT2, 11: INR2, 12: INR2}.get(col, "#,##0")
        c.border = BOX
        if col == 6:
            c.font = LBL
m.conditional_formatting.add("F5:F28", CellIsRule(operator="greaterThan", formula=["0"], fill=GOOD))
m.conditional_formatting.add("F5:F28", CellIsRule(operator="lessThan", formula=["0"], fill=BAD))
m["A31"] = "Return on margin here is the month's NET divided by that month's AVERAGE margin blocked — comparable across months even as lots change."
m["A31"].font = NOTE

# ════════════════════════════════════════════════════════ 4-6. CHARTS
def chart_sheet(title, subtitle):
    c = wb.create_sheet(title)
    c.sheet_view.showGridLines = False
    c["A1"] = title.replace("Chart · ", "")
    c["A1"].font = H1
    c["A2"] = subtitle
    c["A2"].font = NOTE
    return c

# ---- Equity curve + drawdown
e = chart_sheet("Chart · Equity", "Cumulative NET and peak-to-trough drawdown. The shape matters more than the level at n=7.")
for i, h in enumerate(["Date", "Daily NET", "Cumulative", "Peak", "Drawdown"], start=1):
    cc = e.cell(row=4, column=i, value=h)
    cc.font, cc.fill = H2, HD_FILL
    e.column_dimensions[get_column_letter(i)].width = 13
for i in range(NROWS):
    r = 5 + i
    dr = 2 + i
    e.cell(row=r, column=1, value=f"={D}$A${dr}").number_format = DATEF
    e.cell(row=r, column=2, value=f"={D}${C['net_pnl']}${dr}").number_format = INR2
    e.cell(row=r, column=3, value=f"=B{r}" if i == 0 else f"=C{r-1}+B{r}").number_format = INR2
    e.cell(row=r, column=4, value=f"=C{r}" if i == 0 else f"=MAX(D{r-1},C{r})").number_format = INR2
    e.cell(row=r, column=5, value=f"=C{r}-D{r}").number_format = INR2
end = 4 + NROWS
ch = LineChart(); ch.title = "Cumulative NET P&L (₹)"; ch.height, ch.width = 9, 22
ch.add_data(Reference(e, min_col=3, min_row=4, max_row=end), titles_from_data=True)
ch.set_categories(Reference(e, min_col=1, min_row=5, max_row=end))
ch.y_axis.title = "₹ cumulative"; ch.x_axis.title = "trading day"
e.add_chart(ch, "G4")
ch2 = BarChart(); ch2.title = "Drawdown from peak (₹)"; ch2.height, ch2.width = 8, 22
ch2.add_data(Reference(e, min_col=5, min_row=4, max_row=end), titles_from_data=True)
ch2.set_categories(Reference(e, min_col=1, min_row=5, max_row=end))
e.add_chart(ch2, "G23")

# ---- Daily net + what the day offered
d = chart_sheet("Chart · Daily", "Daily NET, and how it compares with the day's best (MFE) and worst (MAE) marks. MFE/MAE are GROSS and sample-based — see Notes.")
for i, h in enumerate(["Date", "NET", "MFE (gross)", "MAE (gross)", "DTE"], start=1):
    cc = d.cell(row=4, column=i, value=h)
    cc.font, cc.fill = H2, HD_FILL
    d.column_dimensions[get_column_letter(i)].width = 13
for i in range(NROWS):
    r, dr = 5 + i, 2 + i
    d.cell(row=r, column=1, value=f"={D}$A${dr}").number_format = DATEF
    d.cell(row=r, column=2, value=f"={D}${C['net_pnl']}${dr}").number_format = INR2
    d.cell(row=r, column=3, value=f"={D}${C['mfe']}${dr}").number_format = INR2
    d.cell(row=r, column=4, value=f"={D}${C['mae']}${dr}").number_format = INR2
    d.cell(row=r, column=5, value=f"={D}${C['dte']}${dr}").number_format = "0"
b = BarChart(); b.type = "col"; b.title = "Daily NET P&L (₹)"; b.height, b.width = 9, 22
b.add_data(Reference(d, min_col=2, min_row=4, max_row=end), titles_from_data=True)
b.set_categories(Reference(d, min_col=1, min_row=5, max_row=end))
d.add_chart(b, "G4")
b2 = BarChart(); b2.type = "col"; b2.grouping = "clustered"
b2.title = "What the day offered vs what we took"; b2.height, b2.width = 9, 22
b2.add_data(Reference(d, min_col=2, max_col=4, min_row=4, max_row=end), titles_from_data=True)
b2.set_categories(Reference(d, min_col=1, min_row=5, max_row=end))
d.add_chart(b2, "G23")

# ---- Costs and DTE behaviour
k = chart_sheet("Chart · Costs & DTE", "Where the money goes, and whether days-to-expiry changes the result. Both are open questions at n=7.")
for i, h in enumerate(["Date", "Gross", "Charges", "Slippage", "DTE", "Return on margin"], start=1):
    cc = k.cell(row=4, column=i, value=h)
    cc.font, cc.fill = H2, HD_FILL
    k.column_dimensions[get_column_letter(i)].width = 14
for i in range(NROWS):
    r, dr = 5 + i, 2 + i
    k.cell(row=r, column=1, value=f"={D}$A${dr}").number_format = DATEF
    k.cell(row=r, column=2, value=f"={D}${C['gross_pnl']}${dr}").number_format = INR2
    k.cell(row=r, column=3, value=f"={D}${C['charges']}${dr}").number_format = INR2
    k.cell(row=r, column=4, value=f"=N({D}${C['slip_entry']}${dr})+N({D}${C['slip_exit']}${dr})").number_format = INR2
    k.cell(row=r, column=5, value=f"={D}${C['dte']}${dr}").number_format = "0"
    k.cell(row=r, column=6, value=f"={D}${C['roi_on_margin_pct']}${dr}/100").number_format = PCT3
c3 = BarChart(); c3.type = "col"; c3.grouping = "clustered"
c3.title = "Gross vs charges vs slippage (₹)"; c3.height, c3.width = 9, 22
c3.add_data(Reference(k, min_col=2, max_col=4, min_row=4, max_row=end), titles_from_data=True)
c3.set_categories(Reference(k, min_col=1, min_row=5, max_row=end))
k.add_chart(c3, "H4")
sc = ScatterChart(); sc.title = "Return on margin vs DTE"; sc.height, sc.width = 9, 22
sc.x_axis.title = "DTE at entry"; sc.y_axis.title = "return on margin"
ser = Series(Reference(k, min_col=6, min_row=5, max_row=end),
             Reference(k, min_col=5, min_row=5, max_row=end), title="day")
ser.marker = Marker(symbol="circle", size=8); ser.graphicalProperties.line.noFill = True
sc.series.append(ser)
k.add_chart(sc, "H23")

# ════════════════════════════════════════════════════════ 7. PROJECTION
p = wb.create_sheet("Projection")
p.sheet_view.showGridLines = False
p["A1"] = "Projection — anchored on the MEASURED pace, tunable"
p["A1"].font = H1
p["A2"] = ("Unlike straddle-income-growth-analysis.xlsx (a pure calculator), this starts from what we have "
           "actually achieved. Yellow cells are yours to tune; the grey 'measured' column is what the live "
           "data says, so you can always see how far a tuned assumption departs from evidence.")
p["A2"].font = NOTE
p["A2"].alignment = Alignment(wrap_text=True)
p.row_dimensions[2].height = 30

def pin(row, label, value, fmt=None, measured=None, note=None):
    p.cell(row=row, column=1, value=label).font = LBL
    c = p.cell(row=row, column=2, value=value)
    c.fill, c.border = IN_FILL, BOX
    if fmt:
        c.number_format = fmt
    if measured is not None:
        mc = p.cell(row=row, column=3, value=measured)
        mc.font = Font(size=9, color="777777")
        if fmt:
            mc.number_format = fmt
    if note:
        p.cell(row=row, column=4, value=note).font = NOTE
    return f"$B${row}"

def psec(row, text):
    p.cell(row=row, column=1, value=text).font = Font(bold=True, color="1F3864")
    for c in range(1, 6):
        p.cell(row=row, column=c).fill = SEC_FILL

p["C4"] = "measured"
p["C4"].font = Font(bold=True, size=9, color="777777")
psec(4, "INPUTS")
_corpus_f = f"={OPENING_CASH}+SUM({rng('net_pnl')})"
n_corpus = pin(5, "Starting corpus (INR)", None, INR, _corpus_f,
               "grey = opening cash + every net_pnl since; refreshes itself as rows land"
               if OPENING_CASH else
               "⚠ OPENING_CASH not set — this is CUMULATIVE P&L ONLY, not your balance")
p["B5"] = _corpus_f
if not OPENING_CASH:
    # the note cell already carries the warning text — just make it impossible to skim past
    p["D5"].font = Font(bold=True, color="C00000")
n_lotv = pin(6, "Margin per lot (INR)", None, INR,
             "=ROUND('Live Status'!B16,0)", "grey = measured from live fills")
p[f"B6"] = "='Live Status'!B16"
n_roi = pin(7, "Monthly NET return on margin", None, PCT2,
            "='Live Status'!B21", "grey = our naive run-rate. TUNE THIS — it is the whole model")
p["B7"] = "='Live Status'!B21"
n_months = pin(8, "Months to project", 36, "0", None, f"1 to {PROJ_MONTHS}")
n_grow = pin(9, "Margin per lot — growth p.a.", 0.05, PCT2, None,
             "margin per lot rises with NIFTY/premium levels")
n_reinv = pin(10, "Reinvest profits into more lots?", "Yes", None, None,
              "No = fixed lot count, profits accumulate as idle cash")
n_wpct = pin(11, "Annual withdrawal — % of profit", 0.10, PCT2)
n_wcap = pin(12, "Annual withdrawal — cap (INR)", 10000000, INR, None, "whichever is LOWER")
n_tax = pin(13, "Income tax — effective rate", 0.3120, PCT2, None,
            "New regime incl. 30% base + surcharge + cess. 31.20% up to ₹50L, 34.32% to ₹1cr, 35.88% to ₹2cr, 39% above")
p["A15"] = ("⚠  The monthly return is the ONE input that decides everything, and at n=7 days our 95% "
            "confidence band on it runs from roughly −7% to +18% a month. Treat any single projection as "
            "one draw from a very wide distribution, not a plan.")
p["A15"].font = Font(italic=True, size=9, color="C00000")
p["A15"].alignment = Alignment(wrap_text=True)
p.row_dimensions[15].height = 28
for col, w in (("A", 34), ("B", 18), ("C", 16), ("D", 60), ("E", 12)):
    p.column_dimensions[col].width = w
dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
p.add_data_validation(dv); dv.add(p["B10"])

psec(17, "PROJECTION")
PH = ["Month #", "Month", "Margin/lot", "Opening corpus", "Lots", "Deployed", "NET profit",
      "Income tax", "Withdrawal", "Closing corpus", "Cumulative withdrawn"]
for i, h in enumerate(PH, start=1):
    c = p.cell(row=18, column=i, value=h)
    c.font, c.fill, c.border = H2, HD_FILL, BOX
    c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    p.column_dimensions[get_column_letter(i)].width = 15 if i > 2 else 11
for i in range(PROJ_MONTHS):
    r = 19 + i
    prev = r - 1
    # NOTE the leading '=' — without it Excel stores the whole thing as literal TEXT and the
    # sheet silently renders as strings. Exactly the defect that nearly shipped in
    # straddle-income-growth-analysis.xlsx; caught there and here only by recalculating.
    g = f'=IF($A{r}>$B$8,"",'
    f = {
        1: f"={i+1}",
        2: f'{g}EOMONTH(TODAY(),$A{r}-1))',
        3: f'{g}$B$6*(1+$B$9)^(($A{r}-1)/12))',
        4: f'{g}IF($A{r}=1,$B$5,$J{prev}))',
        # Reinvest=No pins the lot count to what the starting corpus supports
        5: f'{g}IF($B$10="Yes",INT($D{r}/$C{r}),INT($B$5/$C{r})))',
        6: f'{g}$E{r}*$C{r})',
        7: f'{g}$F{r}*$B$7)',
        8: f'{g}MAX(0,$G{r})*$B$13)',
        # withdrawal once a year, on post-tax profit of the trailing 12 months
        9: f'{g}IF(MOD($A{r},12)=0,MIN($B$11*MAX(0,SUM(OFFSET($G{r},-11,0,12,1))-SUM(OFFSET($H{r},-11,0,12,1))),$B$12),0))',
        10: f'{g}$D{r}+$G{r}-$H{r}-$I{r})',
        11: f'{g}IF($A{r}=1,$I{r},$K{prev}+$I{r}))',
    }
    for col, formula in f.items():
        c = p.cell(row=r, column=col, value=formula)
        c.number_format = {2: 'mmm-yyyy', 3: INR, 4: INR, 6: INR, 7: INR,
                           8: INR, 9: INR, 10: INR, 11: INR}.get(col, "#,##0")
        c.border = BOX
        if col in (5, 10):
            c.font = LBL

psec(PROJ_MONTHS + 21, "SENSITIVITY — corpus at the projected horizon, by monthly return")
srow = PROJ_MONTHS + 22
p.cell(row=srow, column=1, value="Monthly return").font = H2
p.cell(row=srow, column=1).fill = HD_FILL
for j, lab in enumerate(["→ closing corpus", "→ total withdrawn", "→ corpus + withdrawn", "vs starting"], start=2):
    c = p.cell(row=srow, column=j, value=lab)
    c.font, c.fill = H2, HD_FILL
p.cell(row=srow + 1, column=1, value="(this sheet's input)").font = NOTE
p.cell(row=srow + 1, column=2, value=f"=INDEX($J$19:$J${18+PROJ_MONTHS},$B$8)").number_format = INR
p.cell(row=srow + 1, column=3, value=f"=INDEX($K$19:$K${18+PROJ_MONTHS},$B$8)").number_format = INR
p.cell(row=srow + 1, column=4, value=f"=B{srow+1}+C{srow+1}").number_format = INR
p.cell(row=srow + 1, column=5, value=f"=B{srow+1}/$B$5").number_format = '0.0"×"'
p.cell(row=srow + 3, column=1,
       value="For other rates, change B7 — the table above recomputes. A separate scenario grid would need "
             "circular references, so this deliberately reads the live model rather than duplicating it.").font = NOTE

# ════════════════════════════════════════════════════════ 8. NOTES
n = wb.create_sheet("Notes")
n.sheet_view.showGridLines = False
n["A1"] = "How to read this workbook"
n["A1"].font = H1
TXT = [
    ("Source of truth", "Every sheet derives from log/trade_journal.csv, which is written automatically at "
                        "15:22 IST by openalgo-trade-journal.timer on the server. Pull it down with "
                        "strategies/scripts/sync_from_server.sh, then re-run "
                        "strategies/scripts/trade_analytics_xlsx.py to rebuild this file."),
    ("Data sheet is a real Excel Table", "It is named `Journal`, so the filter dropdowns work and you can sort/filter "
                                         "freely. Everything else is formulas pointing at it, so filtering the view "
                                         "does NOT change the Summary — that is deliberate; totals should not move "
                                         "because you filtered."),
    ("Adding rows", "Formulas span rows 2:400. Paste new journal rows under the existing ones and Live Status, "
                    "Monthly, the charts and the Projection anchor all update."),
    ("", ""),
    ("THE TWO MARGIN NUMBERS", "`margin_blocked` is what the broker actually blocked; `max_risk_defined` is the "
                               "position's worst case if the index runs past a wing. They differ by ~4.6× and are NOT "
                               "interchangeable. Every return figure here uses margin_blocked. The Telegram alerts had "
                               "this wrong until 2026-08-17 and it inflated return-on-margin ~5× on winning days."),
    ("MFE / MAE are GROSS", "They come from the strategy's own ~5-second marks, not from fills, so they carry no "
                            "charges and no spread. Use them to see what a day OFFERED, never to reconcile to net_pnl. "
                            "2026-08-13 peaked at +1,443 gross and closed −416 net."),
    ("Slippage is already inside gross", "It is shown to size the cost, not to subtract again. Currently it runs at "
                                         "roughly a third of net profit and is concentrated at the EXIT — the entry "
                                         "figure is a conservative lower bound (measured post-fill), the exit figure "
                                         "is exact."),
    ("Confidence column", "high = every fill price recovered and gross reconciles to the broker's m2mrealized. "
                          "medium/low rows are reconstructions and should be excluded from fill-level analysis."),
    ("", ""),
    ("WHY THE t-STATISTIC IS ON THE SUMMARY", "A short-vol strategy wins most days and loses rarely, so a handful of "
                                              "good days can look exactly like an edge. Until |t| ≈ 2 the return is "
                                              "not distinguishable from zero. Watch the 'top-2 days share of net' "
                                              "line alongside it: if two days carry the whole result, it is a tail, "
                                              "not a trend."),
    ("Where the starting corpus comes from", "Projection!B5 is a formula: opening cash + SUM of "
                                             "every net_pnl in Data. It therefore tracks the real "
                                             "account balance and never needs hand-editing. The "
                                             "opening figure is read from the OPENING_CASH env var "
                                             "or log/opening_cash.txt — deliberately NOT in the "
                                             "source, because the git remote is a public fork and "
                                             "an opening balance is an account size. If it is "
                                             "unset the cell shows cumulative P&L only and says so "
                                             "in red."),
    ("Projection vs the growth calculator", "straddle-income-growth-analysis.xlsx answers 'if I earned X% a month, "
                                            "where would I end up?'. This Projection answers 'if the pace we have "
                                            "actually measured continues, where would we end up?' — and lets you tune "
                                            "from there. The measured column beside each input is the anchor; the "
                                            "further your input sits from it, the more you are assuming."),
    ("What the Projection does NOT model", "No losing months, no margin-availability ceiling, no liquidity limit, no "
                                           "change in lot size or exchange margin rules, and no variance at all — it "
                                           "compounds a constant rate. Real results will not be smooth. Treat it as an "
                                           "envelope, not a forecast."),
]
r = 3
for a, b in TXT:
    if a and not b:
        pass
    elif a:
        n.cell(row=r, column=1, value=a).font = LBL
        n.cell(row=r, column=1).alignment = Alignment(vertical="top", wrap_text=True)
        c = n.cell(row=r, column=2, value=b)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
n.column_dimensions["A"].width = 38
n.column_dimensions["B"].width = 105

wb.save(OUT)
print(f"written: {OUT}")
print(f"sheets : {wb.sheetnames}")
print(f"data   : {NROWS} journal rows, {len(COLS)} columns")
