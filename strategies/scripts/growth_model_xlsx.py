#!/usr/bin/env python3
"""Build straddle-income-growth-analysis.xlsx — a LIVE-FORMULA model (no baked numbers).

Every output cell is an Excel formula reading named input cells, so changing any input on
the Inputs sheet recomputes the whole workbook. Nothing is precomputed in Python.

Model, per month, per scenario:
    lots       = INT(corpus / lot_value)          <- reinvest only when a WHOLE lot is affordable
    deployed   = lots * lot_value                  <- idle remainder earns NOTHING
    profit     = deployed * monthly_return
    corpus    += profit - tax(quarter end) - withdrawal(year end)
    lot_value  = lot_value_0 * (1 + growth) ^ (year - 1)
"""
import datetime as dt

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import os
REPO = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", ".."))
OUT = os.getenv("XLSX_OUT", os.path.join(REPO, "log", "straddle-income-growth-analysis.xlsx"))
MAXY = int(os.getenv("XLSX_MAXY", "30"))          # physical rows built; the Years input blanks out anything beyond it
HDR = 4            # engine header row -> month m lives on row HDR + m

INR = '[>=10000000]₹##\\,##\\,##\\,##0;[>=100000]₹##\\,##\\,##0;₹##,##0'
INR2 = '[>=10000000]₹##\\,##\\,##\\,##0.00;[>=100000]₹##\\,##\\,##0.00;₹##,##0.00'
PCT2 = '0.00%'
DATEF = 'mmm-yyyy'

H1 = Font(bold=True, size=14, color="1F3864")
H2 = Font(bold=True, size=11, color="FFFFFF")
LBL = Font(bold=True)
NOTE = Font(italic=True, size=9, color="666666")
IN_FILL = PatternFill("solid", fgColor="FFF2CC")     # yellow = editable input
HD_FILL = PatternFill("solid", fgColor="4472C4")
SEC_FILL = PatternFill("solid", fgColor="D9E1F2")
thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ─────────────────────────────────────────────────────────────── INPUTS
ws = wb.active
ws.title = "Inputs"
ws.sheet_view.showGridLines = False
ws["A1"] = "Straddle Income — Investment Growth Analysis"
ws["A1"].font = H1
ws["A2"] = "Yellow cells are inputs. Everything else is a live formula — change any input and the whole workbook recomputes."
ws["A2"].font = NOTE


def section(row, text):
    ws.cell(row=row, column=1, value=text).font = Font(bold=True, color="1F3864")
    for c in range(1, 7):
        ws.cell(row=row, column=c).fill = SEC_FILL


def inp(row, label, value, fmt=None, note=None):
    ws.cell(row=row, column=1, value=label).font = LBL
    c = ws.cell(row=row, column=2, value=value)
    c.fill, c.border = IN_FILL, BOX
    if fmt:
        c.number_format = fmt
    if note:
        ws.cell(row=row, column=3, value=note).font = NOTE
    return f"Inputs!$B${row}"


section(4, "1 · CAPITAL & TIME")
n_start = inp(5, "Starting month", dt.datetime(2026, 9, 1), DATEF, "First month of deployment (Mon-YYYY)")
n_init = inp(6, "Initial investment (INR)", 500000, INR, "Corpus at the starting month")
n_lot0 = inp(7, "Value of 1 lot at start (INR)", 83333, INR, "₹5,00,000 / 6 lots ≈ ₹83,333")
n_grow = inp(8, "Lot value increase — year on year", 0.05, PCT2, "Margin per lot rises as NIFTY/premium grow")
n_yrs = inp(9, "Number of years to model", int(os.getenv("XLSX_YEARS", "10")), "0", f"1 to {MAXY}")
ws["C10"] = "→ Starting lots (derived):"
ws["C10"].font = LBL
ws["E10"] = f"=INT({n_init}/{n_lot0})"
ws["E10"].font = Font(bold=True, color="1F3864")

section(12, "2 · MONTHLY RETURN SCENARIOS  —  enter returns NET of brokerage, STT and slippage")
n_ra = inp(13, "Scenario A — monthly return", 0.05, PCT2)
n_rb = inp(14, "Scenario B — monthly return", 0.08, PCT2)
n_rc = inp(15, "Scenario C — monthly return", 0.10, PCT2)
n_rd = inp(16, "Scenario D — monthly return (your own)", 0.0425, PCT2, "2-decimal precision, e.g. 4.25%")
ws["A17"] = ("Earned on DEPLOYED capital only (lots × lot value); idle cash earns nothing. These % are assumed "
             "ALREADY NET of transaction costs — the only further deductions are govt tax and withdrawals.")
ws["A17"].font = NOTE

section(18, "3 · ANNUAL WITHDRAWAL")
n_wpct = inp(19, "Withdrawal — % of that year's profit", 0.10, PCT2)
n_wcap = inp(20, "Withdrawal — absolute cap (INR)", 10000000, INR, "₹1 crore")
n_wbase = inp(21, "Withdrawal calculated on", "Post-tax profit", None, "Whichever is LOWER of the % and the cap is withdrawn")
ws["A22"] = "Withdrawn each year = MIN( % × profit , cap )"
ws["A22"].font = NOTE

section(24, "4 · GOVERNMENT TAX  (New Regime — 30% base slab already included)")
n_tfreq = inp(25, "Tax deducted", "Quarterly", None, "Quarterly = advance tax leaves the corpus 4×/yr, so it compounds less")
ws["A27"] = "Income range"
ws["B27"] = "Lower bound (INR)"
ws["C27"] = "Surcharge"
ws["D27"] = "Effective rate — NEW  ← USED"
ws["E27"] = "Effective rate — OLD (reference only)"
for c in range(1, 6):
    x = ws.cell(row=27, column=c)
    x.font, x.fill, x.border = H2, HD_FILL, BOX
    x.alignment = Alignment(wrap_text=True, vertical="center")
TAX = [
    ("Up to ₹50 Lakh", 0, 0.00, 0.3120, 0.3120),
    ("₹50 Lakh – ₹1 Crore", 5000000, 0.10, 0.3432, 0.3432),
    ("₹1 Crore – ₹2 Crore", 10000000, 0.15, 0.3588, 0.3588),
    ("₹2 Crore – ₹5 Crore", 20000000, 0.25, 0.3900, 0.3900),
    ("Above ₹5 Crore", 50000000, 0.25, 0.3900, 0.42744),
]
for i, (lbl, lo, sur, new, old) in enumerate(TAX):
    r = 28 + i
    ws.cell(row=r, column=1, value=lbl).border = BOX
    for col, val, fmt in ((2, lo, INR), (3, sur, "0%"), (4, new, "0.000%"), (5, old, "0.000%")):
        c = ws.cell(row=r, column=col, value=val)
        c.number_format, c.border = fmt, BOX
        if col in (2, 4):
            c.fill = IN_FILL
ws["A34"] = ("Slab is chosen from ANNUALISED profit, so a growing book climbs into higher surcharge bands "
             "over time. Only the NEW-regime column drives the maths.")
ws["A34"].font = NOTE

n_tlow, n_trate = "Inputs!$B$28:$B$32", "Inputs!$D$28:$D$32"

for col, w in zip("ABCDEF", (38, 20, 30, 30, 30, 14)):
    ws.column_dimensions[col].width = w

dv1 = DataValidation(type="list", formula1='"Post-tax profit,Gross profit"', allow_blank=False)
ws.add_data_validation(dv1)
dv1.add(ws["B21"])
dv2 = DataValidation(type="list", formula1='"Quarterly,Annually"', allow_blank=False)
ws.add_data_validation(dv2)
dv2.add(ws["B25"])

names = {"StartMonth": n_start, "InitInvest": n_init, "LotValue0": n_lot0, "LotGrowth": n_grow,
         "NYears": n_yrs, "RateA": n_ra, "RateB": n_rb, "RateC": n_rc, "RateD": n_rd,
         "WdrPct": n_wpct, "WdrCap": n_wcap, "WdrBase": n_wbase, "TaxFreq": n_tfreq,
         "TaxLower": n_tlow, "TaxRate": n_trate}
for k, v in names.items():
    wb.defined_names.add(__import__("openpyxl").workbook.defined_name.DefinedName(k, attr_text=v))

# ─────────────────────────────────────────────────────────── ENGINE SHEETS
ENG_COLS = [
    ("Month #", 8, "0"), ("Month", 11, DATEF), ("Yr", 5, "0"), ("M-in-yr", 8, "0"),
    ("Lot value", 14, INR), ("Opening corpus", 16, INR), ("Lots", 7, "0"),
    ("Deployed capital", 16, INR), ("Idle cash", 13, INR), ("Profit (month)", 15, INR),
    ("Profit YTD", 15, INR), ("Profit this qtr", 15, INR), ("Slab income", 15, INR),
    ("Tax rate", 10, "0.000%"), ("Tax deducted", 14, INR), ("Tax YTD", 14, INR),
    ("Withdrawal", 14, INR), ("Closing corpus", 17, INR), ("Lots at close", 12, "0"),
]
SCEN = [("Engine A", "RateA", "A"), ("Engine B", "RateB", "B"),
        ("Engine C", "RateC", "C"), ("Engine D", "RateD", "D")]

for title, rate, tag in SCEN:
    e = wb.create_sheet(title)
    e.sheet_view.showGridLines = False
    e["A1"] = f"Monthly engine — Scenario {tag}"
    e["A1"].font = H1
    e["A2"] = "Monthly return:"
    e["A2"].font = LBL
    e["B2"] = f"={rate}"
    e["B2"].number_format = PCT2
    e["B2"].font = Font(bold=True, color="C00000")
    e["D2"] = "One row per month. Returns accrue on DEPLOYED capital only; idle cash earns nothing until it funds a whole lot."
    e["D2"].font = NOTE
    for i, (h, w, _f) in enumerate(ENG_COLS, start=1):
        c = e.cell(row=HDR, column=i, value=h)
        c.font, c.fill, c.border = H2, HD_FILL, BOX
        c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        e.column_dimensions[get_column_letter(i)].width = w
    e.freeze_panes = f"A{HDR+1}"

    for m in range(1, MAXY * 12 + 1):
        r = HDR + m
        p = r - 1                                     # previous month's row
        g = f'=IF($A{r}>NYears*12,"",'                # NOTE the leading '=' — without it Excel
        #                                              treats the whole cell as literal text
        f = {
            1: f"={m}",
            2: f"{g}EDATE(StartMonth,$A{r}-1))",
            3: f"{g}INT(($A{r}-1)/12)+1)",
            4: f"{g}MOD($A{r}-1,12)+1)",
            5: f"{g}LotValue0*(1+LotGrowth)^($C{r}-1))",
            6: f"{g}IF($A{r}=1,InitInvest,$R{p}))",
            7: f"{g}INT($F{r}/$E{r}))",
            8: f"{g}$G{r}*$E{r})",
            9: f"{g}$F{r}-$H{r})",
            10: f"{g}$H{r}*{rate})",
            # YTD resets every January of the model year
            11: f"{g}IF($D{r}=1,$J{r},$K{p}+$J{r}))",
            12: f"{g}IF(MOD($D{r},3)=0,SUM($J{r-2}:$J{r}),0))",
            # Quarterly: annualise YTD so the slab reflects the run-rate. Annual: use YTD as-is.
            13: f'{g}IF(TaxFreq="Quarterly",$K{r}*12/$D{r},$K{r}))',
            14: f"{g}LOOKUP($M{r},TaxLower,TaxRate))",
            15: (f'{g}IF(TaxFreq="Quarterly",IF(MOD($D{r},3)=0,$L{r}*$N{r},0),'
                 f"IF($D{r}=12,$K{r}*$N{r},0)))"),
            16: f"{g}IF($D{r}=1,$O{r},$P{p}+$O{r}))",
            17: (f'{g}IF($D{r}=12,MIN(WdrPct*IF(WdrBase="Post-tax profit",$K{r}-$P{r},$K{r}),'
                 f"WdrCap),0))"),
            18: f"{g}$F{r}+$J{r}-$O{r}-$Q{r})",
            19: f"{g}INT($R{r}/$E{r}))",
        }
        for col, formula in f.items():
            c = e.cell(row=r, column=col, value=formula)
            c.number_format = ENG_COLS[col - 1][2]
            if m % 12 == 0:                            # tint the year-end row
                c.fill = PatternFill("solid", fgColor="EDEDED")

# ─────────────────────────────────────────────────────────────── SUMMARY
s = wb.create_sheet("Summary", 1)
s.sheet_view.showGridLines = False
s["A1"] = "Annual Summary — one row per year, per scenario"
s["A1"].font = H1
s["A2"] = '=TEXT(StartMonth,"mmm-yyyy")&"  to  "&TEXT(EDATE(StartMonth,NYears*12-1),"mmm-yyyy")&"   ·   "&TEXT(NYears,"0")&" years   ·   initial "&TEXT(InitInvest,"₹##\\,##\\,##0")&" = "&TEXT(INT(InitInvest/LotValue0),"0")&" lots"'
s["A2"].font = Font(italic=True, color="1F3864")

s["A4"] = "HEADLINE — where each scenario ends up"
s["A4"].font = Font(bold=True, color="1F3864")
for c in range(1, 9):
    s.cell(row=4, column=c).fill = SEC_FILL
HH = ["Scenario", "Monthly return", "Final corpus", "Final lots", "Lots growth ×",
      "Total tax paid", "Total withdrawn", "Corpus ×"]
for i, h in enumerate(HH, start=1):
    c = s.cell(row=5, column=i, value=h)
    c.font, c.fill, c.border = H2, HD_FILL, BOX
    c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

BLK = {}
row = 12
for title, rate, tag in SCEN:
    BLK[tag] = row
    row += MAXY + 3

SUM_COLS = [("Year", 6, "0"), ("Year ending", 12, DATEF), ("Lot value", 13, INR),
            ("Opening corpus", 16, INR), ("Lots start", 10, "0"), ("Lots end", 9, "0"),
            ("Lots added", 10, "0"), ("Deployed capital", 16, INR),
            ("TRADING PROFIT\n(net of charges, pre-tax)", 20, INR), ("Govt tax", 15, INR), ("Withdrawal", 15, INR),
            ("REAL PROFIT (post tax & withdrawal)", 20, INR), ("Closing corpus", 17, INR)]

for title, rate, tag in SCEN:
    top = BLK[tag]
    s.cell(row=top, column=1, value=f"SCENARIO {tag}").font = Font(bold=True, size=12, color="1F3864")
    s.cell(row=top, column=2, value=f"={rate}").number_format = PCT2
    s.cell(row=top, column=2).font = Font(bold=True, color="C00000")
    s.cell(row=top, column=3, value="per month").font = NOTE
    for c in range(1, len(SUM_COLS) + 1):
        s.cell(row=top, column=c).fill = SEC_FILL
    for i, (h, w, _f) in enumerate(SUM_COLS, start=1):
        c = s.cell(row=top + 1, column=i, value=h)
        c.font, c.fill, c.border = H2, HD_FILL, BOX
        c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        s.column_dimensions[get_column_letter(i)].width = w
    for y in range(1, MAXY + 1):
        r = top + 1 + y
        last = HDR + y * 12                 # engine row for that year's LAST month
        first = HDR + (y - 1) * 12 + 1      # engine row for that year's FIRST month
        g = f'=IF($A{r}>NYears,"",'
        f = {
            1: f"={y}",
            2: f"{g}EDATE(StartMonth,$A{r}*12-1))",
            3: f"{g}'{title}'!$E{last})",
            4: f"{g}'{title}'!$F{first})",
            5: f"{g}'{title}'!$G{first})",
            6: f"{g}'{title}'!$S{last})",
            7: f"{g}'{title}'!$S{last}-'{title}'!$G{first})",
            8: f"{g}'{title}'!$S{last}*'{title}'!$E{last})",
            9: f"{g}'{title}'!$K{last})",
            10: f"{g}'{title}'!$P{last})",
            11: f"{g}'{title}'!$Q{last})",
            12: f"{g}'{title}'!$K{last}-'{title}'!$P{last}-'{title}'!$Q{last})",
            13: f"{g}'{title}'!$R{last})",
        }
        for col, formula in f.items():
            c = s.cell(row=r, column=col, value=formula)
            c.number_format = SUM_COLS[col - 1][2]
            c.border = BOX
            if col in (9, 12, 13):
                c.font = Font(bold=True)

# headline rows reference each block's final modelled year
for i, (title, rate, tag) in enumerate(SCEN):
    r = 6 + i
    top = BLK[tag]
    fin = f"INDEX(${get_column_letter(13)}${top+2}:${get_column_letter(13)}${top+1+MAXY},NYears)"
    lots = f"INDEX($F${top+2}:$F${top+1+MAXY},NYears)"
    s.cell(row=r, column=1, value=f"Scenario {tag}").font = LBL
    s.cell(row=r, column=2, value=f"={rate}").number_format = PCT2
    s.cell(row=r, column=3, value=f"={fin}").number_format = INR
    s.cell(row=r, column=4, value=f"={lots}").number_format = "0"
    s.cell(row=r, column=5, value=f"={lots}/INT(InitInvest/LotValue0)").number_format = '0.0"×"'
    s.cell(row=r, column=6, value=f"=SUM($J${top+2}:$J${top+1+MAXY})").number_format = INR
    s.cell(row=r, column=7, value=f"=SUM($K${top+2}:$K${top+1+MAXY})").number_format = INR
    s.cell(row=r, column=8, value=f"=({fin}+SUM($K${top+2}:$K${top+1+MAXY}))/InitInvest").number_format = '0.0"×"'
    for c in range(1, 9):
        s.cell(row=r, column=c).border = BOX
        if c in (3, 4, 8):
            s.cell(row=r, column=c).font = Font(bold=True, color="1F3864")
s["A10"] = ('"Corpus ×" counts money taken out as well as money still invested — otherwise a big withdrawal '
            "looks like worse performance. Lots growth is the honest measure of deployed scale.")
s["A10"].font = NOTE
s.freeze_panes = "A6"

# ─────────────────────────────────────────────────────────────── NOTES
n = wb.create_sheet("Assumptions")
n.sheet_view.showGridLines = False
n["A1"] = "Assumptions & how to read this model"
n["A1"].font = H1
TEXT = [
    ("", ""),
    ("THE DEDUCTION WATERFALL — what comes off the return you enter", ""),
    ("The % you enter is ALREADY net of transaction costs",
     "Brokerage, STT, exchange/SEBI charges, stamp duty, GST and slippage are assumed to be inside the number "
     "you type. The model does NOT deduct them again. Only two liabilities are taken off it: government tax, "
     "and your annual withdrawal."),
    ("So the chain is",
     "deployed capital × monthly return  →  minus govt tax (quarterly)  →  minus annual withdrawal  →  "
     "remainder compounds. Nothing else is subtracted anywhere in the workbook."),
    ('Why the column says "TRADING PROFIT (net of charges, pre-tax)"',
     'It is deliberately not called "gross" — that word would suggest charges are still to come. They are not. '
     "It is profit after trading costs and before government tax."),
    ("Reality check on entering a number",
     "Live experience at 2 lots has charges running 28–43% of gross trading profit. So if your raw strategy "
     "gross is 8% a month, the figure to enter here is nearer 5%. Entering 8% asserts you already cleared costs."),
    ("", ""),
    ("WHAT DRIVES THE COMPOUNDING", ""),
    ("Returns accrue on DEPLOYED capital only",
     "Profit = (lots × lot value) × monthly return. Cash that is not enough to fund a whole extra lot sits "
     "idle and earns nothing. This is the single most important assumption — it is why the corpus grows in "
     "steps rather than smoothly, and why a higher lot value slows growth."),
    ("Reinvestment is lumpy, by design",
     "Lots = INT(corpus / lot value), recomputed every month. One more lot is deployed only when the whole "
     "margin for it is available, matching how the straddle actually scales."),
    ("Lot value rises every year",
     "Margin per lot grows with NIFTY and premium levels, so the bar for adding a lot keeps rising. Default "
     "5% a year. Set it to 0% to see how much this drag alone costs you."),
    ("", ""),
    ("TAX", ""),
    ("New regime, 30% base slab already inside the rates",
     "The effective rates supplied (31.20% → 39.00%) already contain the 30% base slab plus surcharge and "
     "cess. Only the NEW-regime column is used; the OLD-regime column is kept for reference."),
    ("Slab is picked from ANNUALISED profit",
     'With "Quarterly" selected, the slab is chosen from year-to-date profit scaled to a full year — the way '
     "advance tax is estimated in practice. A growing book therefore climbs into higher surcharge bands over "
     'time. With "Annually", the slab is chosen from the actual full-year profit.'),
    ("Quarterly vs Annually is NOT cosmetic",
     "Quarterly deduction removes money from the corpus four times a year instead of once, so it compounds "
     "less. Switching to Annually will flatter the result. Quarterly is the realistic default."),
    ("", ""),
    ("WITHDRAWAL", ""),
    ("MIN(% of profit, cap), once a year",
     "Default 10% of profit capped at ₹1 crore — whichever is LOWER. In early years the 10% binds; only once "
     "annual profit passes ₹10 crore does the cap start binding."),
    ('Base defaults to "Post-tax profit"',
     "You withdraw from what is actually yours after tax. Switch to \"Gross profit\" on the Inputs sheet if "
     "you intend the 10% to be taken before tax — it withdraws more and slows compounding."),
    ("", ""),
    ("WHAT THIS MODEL DOES NOT DO", ""),
    ("No losing months",
     "A constant positive monthly return is a planning tool, not a forecast. The live straddle has already "
     "had a losing day (13-Aug-2026, −₹686) and its worst backtested day was −₹17,522. Treat every number "
     "here as an upper envelope, not an expectation."),
    ("No margin-availability ceiling",
     "The model will happily deploy hundreds of lots. Real NIFTY option liquidity, exchange position limits "
     "and slippage all bite well before that — and slippage grows with size."),
    ("No brokerage or transaction cost",
     "The monthly return you enter must therefore be NET of brokerage, STT and slippage. Live experience: "
     "charges have been running 28–43% of gross profit at 2 lots, so a gross 8% is nowhere near a net 8%."),
    ("Returns are on capital, not on a strategy",
     "5% a month sustained for 10 years is 1.05^120 ≈ 348× before reinvestment effects. Scenarios that look "
     "modest per month become extraordinary over a decade — which is the point of the exercise, but also the "
     "reason to be sceptical of the top end."),
]
r = 3
for a, b in TEXT:
    if a and not b:
        n.cell(row=r, column=1, value=a).font = Font(bold=True, size=11, color="1F3864")
    elif a:
        n.cell(row=r, column=1, value=a).font = LBL
        n.cell(row=r, column=1).alignment = Alignment(vertical="top", wrap_text=True)
        c = n.cell(row=r, column=2, value=b)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
n.column_dimensions["A"].width = 42
n.column_dimensions["B"].width = 105
for rr in range(3, r):
    n.row_dimensions[rr].height = None

wb.save(OUT)
print(f"written: {OUT}")
print(f"sheets : {wb.sheetnames}")
